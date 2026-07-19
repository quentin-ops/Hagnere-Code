#!/usr/bin/env python3
"""Verify the workbook's decision formulas after a real LibreOffice recalculation.

The source workbook is only read. Every scenario is applied to an isolated
copy, converted by LibreOffice into a second temporary directory, then read
with ``data_only=True`` so assertions use the cached values produced by the
spreadsheet engine rather than formulas interpreted in Python.
"""

from __future__ import annotations

import argparse
import hashlib
import os
import shutil
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.properties import CalcProperties


@dataclass(frozen=True)
class Scenario:
    slug: str
    label: str
    values: dict[str, object]
    expected_decision: str
    expected_summary: str


TEST_DATE = date(2026, 7, 19)
SCENARIOS = (
    Scenario(
        "initial",
        "État initial",
        {"J5": "À tester", "K5": None, "L5": None, "M5": "—", "N5": None, "O5": None},
        "En attente",
        "RECETTE EN COURS",
    ),
    Scenario(
        "conforme-incomplet",
        "Conforme sans date ni preuve",
        {"J5": "Conforme", "K5": None, "L5": None, "M5": "—", "N5": None, "O5": None},
        "INCOMPLET",
        "RECETTE INCOMPLÈTE",
    ),
    Scenario(
        "conforme-sans-date",
        "Conforme sans date",
        {
            "J5": "Conforme",
            "K5": None,
            "L5": "Capture datée QA-001",
            "M5": "—",
            "N5": None,
            "O5": None,
        },
        "INCOMPLET",
        "RECETTE INCOMPLÈTE",
    ),
    Scenario(
        "conforme-sans-preuve",
        "Conforme sans preuve",
        {"J5": "Conforme", "K5": TEST_DATE, "L5": None, "M5": "—", "N5": None, "O5": None},
        "INCOMPLET",
        "RECETTE INCOMPLÈTE",
    ),
    Scenario(
        "conforme-complet",
        "Conforme avec date et preuve",
        {
            "J5": "Conforme",
            "K5": TEST_DATE,
            "L5": "Capture datée QA-001",
            "M5": "—",
            "N5": None,
            "O5": None,
        },
        "OK",
        "RECETTE EN COURS",
    ),
    Scenario(
        "a-corriger-gravite-vide",
        "À corriger avec gravité vide",
        {
            "J5": "À corriger",
            "K5": TEST_DATE,
            "L5": None,
            "M5": None,
            "N5": "Écart QA-002",
            "O5": None,
        },
        "INCOMPLET",
        "RECETTE INCOMPLÈTE",
    ),
    Scenario(
        "a-corriger-gravite-tiret",
        "À corriger avec gravité « — »",
        {
            "J5": "À corriger",
            "K5": TEST_DATE,
            "L5": None,
            "M5": "—",
            "N5": "Écart QA-002",
            "O5": None,
        },
        "INCOMPLET",
        "RECETTE INCOMPLÈTE",
    ),
    Scenario(
        "a-corriger-sans-date",
        "À corriger sans date",
        {
            "J5": "À corriger",
            "K5": None,
            "L5": None,
            "M5": "Majeur",
            "N5": "Écart QA-003",
            "O5": None,
        },
        "INCOMPLET",
        "RECETTE INCOMPLÈTE",
    ),
    Scenario(
        "a-corriger-sans-anomalie",
        "À corriger sans anomalie",
        {"J5": "À corriger", "K5": TEST_DATE, "L5": None, "M5": "Majeur", "N5": None, "O5": None},
        "INCOMPLET",
        "RECETTE INCOMPLÈTE",
    ),
    Scenario(
        "a-corriger-majeur",
        "À corriger Majeur complet",
        {
            "J5": "À corriger",
            "K5": TEST_DATE,
            "L5": None,
            "M5": "Majeur",
            "N5": "Écart QA-003",
            "O5": None,
        },
        "À traiter",
        "RECETTE EN COURS",
    ),
    Scenario(
        "a-corriger-bloquant",
        "À corriger Bloquant complet",
        {
            "J5": "À corriger",
            "K5": TEST_DATE,
            "L5": None,
            "M5": "Bloquant",
            "N5": "Écart QA-004",
            "O5": None,
        },
        "BLOQUANT",
        "RECETTE BLOQUÉE",
    ),
    Scenario(
        "non-applicable-incomplet",
        "Non applicable sans justification",
        {"J5": "Non applicable", "K5": None, "L5": None, "M5": "—", "N5": None, "O5": None},
        "INCOMPLET",
        "RECETTE INCOMPLÈTE",
    ),
    Scenario(
        "non-applicable-justifie",
        "Non applicable avec justification",
        {
            "J5": "Non applicable",
            "K5": None,
            "L5": None,
            "M5": "—",
            "N5": "Hors périmètre contractuel",
            "O5": None,
        },
        "N/A",
        "RECETTE EN COURS",
    ),
)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def resolve_executable(value: str) -> Path:
    candidate = Path(value).expanduser()
    if candidate.exists():
        resolved = candidate.resolve()
    else:
        found = shutil.which(value)
        if not found:
            raise FileNotFoundError(f"LibreOffice/soffice introuvable : {value}")
        resolved = Path(found).resolve()
    if not resolved.is_file() or not os.access(resolved, os.X_OK):
        raise PermissionError(f"LibreOffice/soffice n'est pas exécutable : {resolved}")
    return resolved


def prepare_scenario(source: Path, destination: Path, scenario: Scenario) -> None:
    shutil.copy2(source, destination)
    workbook = load_workbook(destination, data_only=False, keep_links=False)
    try:
        if "Recette" not in workbook.sheetnames or "Synthèse" not in workbook.sheetnames:
            raise ValueError("Feuilles attendues absentes : Recette et/ou Synthèse")
        recipe = workbook["Recette"]
        if not recipe["A5"].value:
            raise ValueError("Recette!A5 doit contenir l'identifiant du test de référence")
        formula = recipe["P5"].value
        if not isinstance(formula, str) or not formula.startswith("="):
            raise ValueError("Recette!P5 ne contient pas une formule")
        for cell, value in scenario.values.items():
            recipe[cell] = value

        # Make the requested recalculation explicit in the OOXML metadata.
        if workbook.calculation is None:
            workbook.calculation = CalcProperties()
        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcOnSave = True
        workbook.calculation.calcCompleted = False
        workbook.save(destination)
    finally:
        workbook.close()


def recalculate_with_libreoffice(
    soffice: Path,
    inputs: list[Path],
    output_dir: Path,
    profile_dir: Path,
    timeout: int,
) -> None:
    command = [
        str(soffice),
        f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
        "--headless",
        "--nologo",
        "--nodefault",
        "--nolockcheck",
        "--nofirststartwizard",
        "--convert-to",
        "xlsx",
        "--outdir",
        str(output_dir),
        *(str(path) for path in inputs),
    ]
    try:
        result = subprocess.run(
            command,
            check=False,
            capture_output=True,
            text=True,
            timeout=timeout,
        )
    except subprocess.TimeoutExpired as exc:
        raise RuntimeError(
            f"LibreOffice n'a pas terminé le recalcul sous {timeout} secondes"
        ) from exc
    if result.returncode != 0:
        details = "\n".join(
            part
            for part in (result.stdout.strip(), result.stderr.strip())
            if part
        )
        raise RuntimeError(
            f"LibreOffice a échoué avec le code {result.returncode}"
            + (f"\n{details}" if details else "")
        )


def read_results(path: Path) -> tuple[object, object, object]:
    workbook = load_workbook(path, data_only=True, read_only=True, keep_links=False)
    try:
        return (
            workbook["Recette"]["P5"].value,
            workbook["Synthèse"]["B5"].value,
            workbook["Synthèse"]["B11"].value,
        )
    finally:
        workbook.close()


def verify(source: Path, soffice: Path, timeout: int) -> None:
    source = source.expanduser().resolve(strict=True)
    if source.suffix.lower() != ".xlsx" or not source.is_file():
        raise ValueError(f"Classeur XLSX attendu : {source}")
    if timeout <= 0:
        raise ValueError("Le délai maximal doit être strictement positif")

    source_digest = sha256(source)
    failures: list[str] = []
    try:
        with tempfile.TemporaryDirectory(prefix="hagnere-xlsx-recalc-") as temp_name:
            temp = Path(temp_name)
            input_dir = temp / "inputs"
            output_dir = temp / "outputs"
            profile_dir = temp / "libreoffice-profile"
            input_dir.mkdir()
            output_dir.mkdir()
            profile_dir.mkdir()

            inputs: list[Path] = []
            for scenario in SCENARIOS:
                scenario_path = input_dir / f"{scenario.slug}.xlsx"
                prepare_scenario(source, scenario_path, scenario)
                inputs.append(scenario_path)

            recalculate_with_libreoffice(
                soffice, inputs, output_dir, profile_dir, timeout
            )

            for scenario in SCENARIOS:
                result_path = output_dir / f"{scenario.slug}.xlsx"
                if not result_path.exists():
                    failures.append(
                        f"{scenario.label}: fichier recalculé absent ({result_path.name})"
                    )
                    continue
                decision, total, summary = read_results(result_path)
                mismatches = []
                if decision != scenario.expected_decision:
                    mismatches.append(
                        f"Recette!P5 attendu {scenario.expected_decision!r}, obtenu {decision!r}"
                    )
                if total != 56:
                    mismatches.append(f"Synthèse!B5 attendu 56, obtenu {total!r}")
                if summary != scenario.expected_summary:
                    mismatches.append(
                        f"Synthèse!B11 attendu {scenario.expected_summary!r}, obtenu {summary!r}"
                    )
                if mismatches:
                    failures.append(f"{scenario.label}: " + "; ".join(mismatches))
                else:
                    print(
                        f"[OK] {scenario.label}: P5={decision!r}, "
                        f"B5={total!r}, B11={summary!r}"
                    )
    finally:
        if not source.exists() or sha256(source) != source_digest:
            raise RuntimeError(
                "Le classeur source a été modifié pendant la vérification"
            )

    if failures:
        formatted = "\n".join(f"- {failure}" for failure in failures)
        raise AssertionError(
            f"Échec de {len(failures)} scénario(s) de recalcul :\n{formatted}"
        )
    print(
        f"Vérification réussie : {len(SCENARIOS)} scénarios recalculés "
        "par LibreOffice ; classeur source inchangé."
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Teste les décisions du XLSX sur des copies recalculées par "
            "LibreOffice. Le classeur source n'est jamais enregistré."
        )
    )
    parser.add_argument("workbook", type=Path, help="classeur XLSX source")
    parser.add_argument(
        "--soffice",
        required=True,
        help="chemin de l'exécutable LibreOffice/soffice, ou nom résolu via PATH",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=180,
        help="délai maximal global du recalcul LibreOffice, en secondes (défaut : 180)",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        verify(args.workbook, resolve_executable(args.soffice), args.timeout)
    except Exception as exc:
        print(f"ERREUR : {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
